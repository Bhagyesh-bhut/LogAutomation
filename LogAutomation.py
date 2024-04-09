from azure.identity import DefaultAzureCredential
from azure.monitor.query import LogsQueryClient, MetricsQueryClient
from datetime import datetime, timezone , timedelta
import os
import pandas as pd
from openpyxl import load_workbook,Workbook
from datetime import datetime,timedelta,date
import queries_list
from azure.storage.blob import BlobServiceClient

today = date.today()
yesterday = today - timedelta(days = 1)

# quries={
#   'alert-count':queries_list.alert_count,
#   'account-count':queries_list.account_count,
#   'disbursements-count':queries_list.disbursement_count,
#   'collateral-count':queries_list.collateral_count,
#   'insurance-count':queries_list.insurance_count,
#   'customer-count':queries_list.customer_count,
#   'customer-error':queries_list.customer_error_count,
#   'alert-error':queries_list.alert_error_count,
#   'account-error':queries_list.account_error_count,
#   'disbursements-error':queries_list.disbursment_error_count,
#   'collateral-error':queries_list.collateral_error_count,
#   'insurance-error':queries_list.insurance_error_count,
# }

quries={
  'alert-count':'exceptions | take 5 | project appName,severityLevel',
  'account-count': 'exceptions | take 1 | project appName,severityLevel'
}

# log_query =  """exceptions | take 5 | project appName,severityLevel """



def create_connection():
  #create azure connection 
  credential = DefaultAzureCredential()
  return LogsQueryClient(credential)
 

def execute_queries(queries,file_name):

  logs_client=create_connection()

  #run the queries one by one
  for key, value in queries.items():
    # print(value)
    sheetName=key
    
    try:
        
      #execte the query in azure app insights logs
      #Please remove timespan parameter from below function if you are using custom timestamp in queries
      response=logs_client.query_resource("subscriptions/c6ac49c8-0c7a-4b98-9d4d-6970b017e7f9/resourceGroups/mule/providers/microsoft.insights/components/azure-app-insight",query=value,timespan=timedelta(days=15))

      # print(response)

      #get the query result
      data=response.tables

      for table in data:
        #convert query result in pandas dataframe
        df = pd.DataFrame(data=table.rows, columns=table.columns)
        # print(df)

      #append result in excel sheet
      export_query_result(file_name,key,df)

    except Exception as e:
      print("An exception occurred")
      print(e)
    
def export_query_result(file_name,sheetName,data):
    
    # Azure Storage account credentials
    # account_name = 'logautomationstorage'
    # account_key = 'qDJSoKxgqoHdcGOEQktNQjDzZGMA7i9wzZ6aW4kJWZ0lg78+u5Ilu/tDuNLIquGaEjX3en7Xe34m+AStcPc0pA=='
    container_name = 'logdata'
    blob_name = file_name  # name of the Excel file in Azure Storage

    # Connect to Azure Storage account
    connection_string ="DefaultEndpointsProtocol=https;AccountName=logautomationstorage;AccountKey=qDJSoKxgqoHdcGOEQktNQjDzZGMA7i9wzZ6aW4kJWZ0lg78+u5Ilu/tDuNLIquGaEjX3en7Xe34m+AStcPc0pA==;EndpointSuffix=core.windows.net"
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    container_client = blob_service_client.get_container_client(container_name)

    # Check if the file exists in Azure Blob Storage
    blob_client = container_client.get_blob_client(blob_name)
    blob_exists = blob_client.exists()

    # Load the workbook if the file exists, otherwise create a new workbook
    if blob_exists:
      # Download the Excel file from Azure Storage
      blob_client = container_client.get_blob_client(blob_name)
      with open(blob_name, "wb") as my_blob:
        download_stream = blob_client.download_blob()
        my_blob.write(download_stream.readall())

      workbook = load_workbook(blob_name)
    else:
      workbook = Workbook()

     
      
    

      #check the sheet name in file if sheet is alerady exits then append the result otherwise create new sheet based on the queries(dictionary) key
    if sheetName in workbook.sheetnames:
      worksheet = workbook[sheetName]
    else:
      worksheet = workbook.create_sheet(title=sheetName)


      #get the column List from the dataframe
    column_list=data.columns.tolist()
      #get the value List from the dataframe
    value_list = data.values.tolist()

      #set the previous day date in sheet because queries run for the collect previous day logs
    worksheet.append([yesterday])

      #append column in sheet
    worksheet.append(column_list)

      #append values in sheet
      #worksheet.append(value_list)
    for row_data in value_list:
      worksheet.append(row_data)

    #save the chnages
    workbook.save(blob_name)

    # Upload the modified Excel file back to Azure Storage, replacing the original file
    with open(blob_name, "rb") as data:
        blob_client.upload_blob(data, overwrite=True)

    print("Data appended successfully.")

# execute_queries(quries,"lock-error-count.xlsx")